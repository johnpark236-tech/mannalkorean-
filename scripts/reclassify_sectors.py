"""
한국 상장 주식 섹터 전면 재분류 S01~S20
입력: D:\logofchoices\data\csv\all_stocks.json
출력: D:\logofchoices\data\csv\stocks_sector_final.csv
      D:\logofchoices\data\csv\stocks_sector_final.xlsx
"""
import json, re
import pandas as pd
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ────────────────────────────────────────────
# 섹터 메타
# ────────────────────────────────────────────
SECTOR_META = {
    'S01': ('반도체·디스플레이·전자',  '반도체, 디스플레이, 전자부품, 컴퓨터·광학기기'),
    'S02': ('바이오·헬스',             '제약, 바이오, 의료기기, 진단, 헬스케어'),
    'S03': ('금융·보험·투자',          '은행, 증권, 보험, 자산운용, VC, 스팩 전체'),
    'S04': ('화학·소재',               '화학, 철강, 비철금속, 플라스틱, 고무, 시멘트, 유리'),
    'S05': ('건설·부동산',             '건설, 토목, 건축, 부동산 개발·임대, 엔지니어링'),
    'S06': ('자동차·부품',             '완성차, 자동차부품, 타이어, 차량용 부품'),
    'S07': ('유통·상사·소매',          '도소매, 종합상사, 백화점, 홈쇼핑, 온라인 유통'),
    'S08': ('조선·방산·우주',          '조선, 선박, 방산, 항공기, 우주·위성'),
    'S09': ('플랫폼·인터넷·소프트웨어','소프트웨어, 플랫폼, 포털, 클라우드, AI, 게임'),
    'S10': ('에너지·자원',             '전력, 가스, 정유, 석유, 광물, 발전'),
    'S11': ('2차전지·신에너지',        '배터리, 소재·장비, 연료전지, 전기차 충전'),
    'S12': ('통신·네트워크·인프라',    '통신서비스, 통신장비, 네트워크, 케이블, 데이터센터'),
    'S13': ('기계·산업장비',           '산업기계, 자동화, 로봇, 측정·제어, 전력기기'),
    'S14': ('미디어·콘텐츠·광고',      '방송, 영화, 음악, 출판, 콘텐츠, 광고'),
    'S15': ('식품·농업·수산',          '식품, 음료, 농업, 비료·농약, 사료, 수산'),
    'S16': ('운송·물류·여행',          '항공, 해운, 육상운송, 물류, 여행, 숙박'),
    'S17': ('생활소비재·패션',         '화장품, 의류, 가구, 생활용품, 레저용품'),
    'S18': ('환경·폐기물·재활용',      '폐기물, 재활용, 수처리, 환경설비'),
    'S19': ('교육·전문서비스',         '교육, 연구, 컨설팅, 시험·인증, 전문서비스'),
    'S20': ('산업재·포장·기타제조',    '금속가공, 포장재, 목재·종이, 범용 산업재'),
}

SECTOR_COLORS = {
    'S01':'DDEEFF','S02':'E2EFDA','S03':'FFF2CC','S04':'FCE4D6','S05':'EDEDED',
    'S06':'DDEBF7','S07':'F2F2F2','S08':'FFD9D9','S09':'C6EFCE','S10':'FFEB9C',
    'S11':'D4F0F0','S12':'E9D9F2','S13':'F9E0C4','S14':'F0D9FF','S15':'D6F0D6',
    'S16':'D9EAD3','S17':'FFE6F0','S18':'D6DCE4','S19':'FFF0D9','S20':'E8E8E8',
}

# ────────────────────────────────────────────
# 우선순위 분류 패턴 (순서 중요)
# ────────────────────────────────────────────
PATTERNS = [

# S03 스팩 최우선
('S03', r'스팩|SPAC|호스팩'),

# S11 2차전지
('S11', r'배터리|이차전지|2차전지|에너지솔루션|에코프로|포스코퓨처|양극재|음극재|전해질|분리막|연료전지|수소연료|충전(?!기계|전기기계|건설)|전지소재|배터리소재|이차전지소재|배터리장비|전지장비|리튬이온|전구체|파우치필름|파우치셀|천보(?!건설)|나노신소재|대주전자재료|신흥에스이씨|SNT에너지|원통형전지|각형전지'),

# S08 조선·방산
('S08', r'조선(?!일보|비즈|닷컴|미디어|방송|신문|뉴스|컴|넷|TV|시대|이코노미|비즈니스)|한화에어로|한화오션|현대중공업|현대미포조선|대한조선|케이조선|대우조선|방산|방위산업|방위(?!비)|항공우주(?!물류)|LIG넥스원|빅텍|퍼스텍|이오시스템|한국항공우주|위성(?!방송콘텐츠)|에어로스페이스|함정|잠수함|전투기|구축함|전차|장갑차|아스트(?!엔터)'),

# S02 바이오·헬스
('S02', r'바이오(?!에너지|가스|디젤|연료|매스|사료|농|수산|종자)|제약|의약|약품(?!유통\b)|신약|임상|항암|항체|세포(?!배양사료)|유전자|단백질|mRNA|백신|진단(?!솔루션SW)|의료기기|의료기|의료AI|헬스케어|헬스(?!클럽)|병원|의원|클리닉(?!IT)|치과|한방|치료(?!시스템IT)|정형|재활의료|원격의료|디지털헬스|건강기능|보톡스|필러(?!반도체)|미용의료|안과|피부과|성형|셀트리온|유한양행|알테오젠|HLB(?!\w)|파마리서치|클래시스|리가켐|루닛|삼천당|동아ST|종근당|한미약품|대웅제약|광동제약|JW중외|보령(?!화학)|일양약품|한독|일동제약|동화약품(?!건설)|부광약품|신풍제약|경보제약|명문제약|태준제약|삼성바이오|녹십자|오스코텍|씨젠|피씨엘|수젠텍|젠큐릭스|비올|메디톡스|휴젤|케어젠|제테마|티앤알바이오|젠(?!트로|킨스|더|코리아가스)|팜(?!레온|레드|어스|디자인)|메드(?!트로닉기계)|큐어(?!콘텐츠)|라이프(?!스타일|코어금속)|이뮨|온코|뉴로|임플란트(?!금속)|덴탈|약학|독성|펩타이드|효소|아미노산|인바디|뷰노|루닛|딥바이오|메디픽셀'),

# S06 자동차·부품
('S06', r'자동차(?!보험|금융|리스|파이낸스|산업공단)|현대차(?!증권)|기아(?!타이거)|모비스|글로비스|만도|한온시스템|HL만도|현대위아|현대모비스|덴소|써미트|세종공업|화신|성우하이텍|에스엘(?!씨|바이오)|대유위니아|인지컨트롤스|서연이화|서연(?!전기)|동원금속|동성화인텍|한국단자|LS오토모티브|KG모빌리티|쌍용차|르노코리아|타이어(?!공업)|한국타이어|넥센타이어|금호타이어|차량부품|자동차부품|차체|헤드램프|ADAS|자율주행(?!SW|소프트)|전동화|차량용|오토바이|이륜차|모토닉|동아오토|파워트레인|트랜스미션|에스엘주식|엠에스오토텍|명화공업|세원정공|드라이브(?!물류)'),

# S01 반도체·디스플레이·전자
('S01', r'반도체|삼성전자|하이닉스|DB하이텍|디스플레이|삼성디스플레이|LG디스플레이|이노텍|자화전자|피에스케이|이오테크닉스|HPSP|한미반도체|레인보우로보틱스|실리콘웍스|엘비세미콘|RFHIC|SFA반도체|동진쎄미켐|솔브레인|케이씨텍|리노공업|퀀타매트릭스|테스(?!나이외)|주성엔지니어링|기가레인|에스앤에스텍|서울반도체|루멘스|에스에프에이|인텍플러스|코스텍시스템|에이피티씨|하나마이크론|두산테스나|웨이퍼|포토레지스트|CMP|에처|CVD|증착|세정|MLCC|커패시터|인덕터|커넥터(?!물류)|LED(?!전선|농업)|광통신부품|포토마스크|블랭크마스크|전자(?!금융|상거래|정부|서류|계약|세금|민원|서명|문서|투표|어음|결제|공시)|정밀광학|광학(?!케미칼|화학|공업)|이미지센서|CMOS|카메라모듈|나노(?!바이오|헬스|환경|농업)|마이크로(?!바이오)|세미콘(?!크리트)|PCB|FPCB|HDI|인쇄회로|기판(?!건설)|갈륨|질화갈륨|탄화규소|SiC|GaN|에피텍시|텔레칩스|어보브반도체|동운아나텍|SK하이닉스|전자소자|액정|OLED|AMOLED|TFT|LCD(?!TV)|광학필름|편광필름|위상차필름|반도체재료|전자재료|수동부품|전자부품(?!유통)'),

# S09 플랫폼·인터넷·소프트웨어
('S09', r'소프트웨어|IT(?!소재|부품|기계|전선|인프라건설)|플랫폼(?!건설)|인터넷(?!통신장비)|포털|클라우드(?!건설)|AI(?!소재|부품)|인공지능|빅데이터|머신러닝|딥러닝|SaaS|ERP|MES|SCM|CRM|솔루션(?!소재|화학|환경|건설|에너지기계)|시스템(?!반도체|에너지설비|건설장비|산업기계|공조기계|기기제조|산업재)|정보기술|IT서비스|IT솔루션|디지털전환|DX(?!전기)|네이버|카카오(?!뱅크|페이)|게임(?!카드|기계)|크래프톤|엔씨소프트|넥슨|펄어비스|넷마블|카카오게임즈|스마일게이트|웹젠|컴투스|위메이드|드래곤플라이|보안(?!시스템건설|기계)|사이버보안|정보보안|핀테크|O2O|모바일(?!통신장비)|메타버스|블록체인|스마트팩토리(?!설비기계)|OTA|테크놀로지(?!화학|기계|자동차)|소프트(?!음료|식품)|데이터(?!소재|화학)|인포(?!메디칼)|닷컴(?!화학)|앱스트리밍|알고리즘|더존비즈온|한글과컴퓨터|웹케시|인크로스|나스미디어|이에이치아이|이루온'),

# S10 에너지·자원
('S10', r'한국전력|한전(?!KDN아이티)|발전(?!기계|설비기계)|전력(?!기기장비|전자부품|반도체)|에너지(?!솔루션배터리|저장배터리|이차전지|전지|바이오매스|솔루션SW)|정유|석유(?!화학)|LNG|LPG|가스(?!반도체장비|에어산업|기계)|원자력|핵연료|원전|광업|광산|자원(?!IT|디지털)|채굴|광물|S-Oil|SK이노베이션|GS칼텍스|현대오일뱅크|태양광(?!장비배터리)|풍력(?!배터리)|수소(?!연료전지배터리)|한국가스공사|도시가스|동아가스|경동도시가스|서울가스|인천도시가스|대성에너지|E1|SK가스|GS에너지|석탄|무연탄|유연탄|포스코에너지|SK에너지'),

# S05 건설·부동산
('S05', r'건설(?!기계|장비|중장비|재료화학|소재화학)|건축(?!소재화학)|토목|엔지니어링(?!소프트|솔루션IT|반도체)|부동산|개발(?!원소프트|소프트|솔루션IT|AI|API)|시행|분양|리츠|리모델링|인테리어(?!IT)|도시개발|주택(?!금융)|GS건설|DL이앤씨|현대건설|대우건설|HDC현대산업개발|롯데건설|SK에코플랜트|포스코건설|한화건설|쌍용건설|태영건설|계룡건설|코오롱글로벌|HL디앤아이한라|토건|재개발|재건축|도시정비|철구|강구조|철탑|지반|기초공사|이앤씨(?!IT)|이엔씨(?!IT)|레미콘(?!화학)|콘크리트(?!화학)|건자재|창호(?!IT)'),

# S07 유통·상사·소매
('S07', r'유통(?!장비|설비기계)|상사(?!법인SPC)|백화점|홈쇼핑|쇼핑(?!몰SW)|면세|아울렛|마트(?!분석AI)|편의점|슈퍼(?!소재)|GS리테일|롯데쇼핑|이마트|신세계|현대홈쇼핑|CJ온스타일|NS홈쇼핑|홈앤쇼핑|무역(?!보험IT)|수출입|종합상사|LG상사|SK네트웍스(?!서비스)|포스코인터내셔널|글로벌(?!기계|화학|바이오|에너지|건설|IT)|인터내셔널(?!에너지)|도매|온라인쇼핑|이커머스|오픈마켓'),

# S12 통신·네트워크·인프라
('S12', r'통신(?!부품소재|반도체|소자)|텔레콤|SK텔레콤|KT(?!\&G|\w*건설|\w*렌탈|\w*부동산)|LG유플러스|LG헬로비전|딜라이브|현대HCN|케이블TV|IPTV|위성방송(?!콘텐츠)|네트워크|5G|LTE|와이파이|중계기(?!의료)|기지국|데이터센터|IDC|CDN|라우터|스위치(?!전자부품)|모뎀|KT스카이라이프|HCN|씨앤앰|안테나(?!의료)|서버호스팅|클라우드인프라'),

# S04 화학·소재
('S04', r'화학(?!바이오|제약|의약)|소재(?!반도체전용|배터리전용|바이오IT)|철강|제철|포스코(?!인터내셔널|건설|에너지)|현대제철|동국제강|세아베스틸|고려아연|풍산(?!홀딩)|LS비철|비철금속|알루미늄|구리|아연|니켈|망간|코발트(?!배터리전용)|시멘트|석회석|규사|유리(?!광학반도체)|석영|플라스틱|합성수지|PE|PP|PET|ABS|에폭시|우레탄|실리콘(?!반도체)|도료|페인트|접착(?!IT)|잉크(?!젯프린터)|불소|염소|황산|인산|NaOH|폴리머|폴리에스터|나일론|섬유(?!IT)|탄소섬유|유리섬유|흑연(?!배터리)|케미칼|케미(?!컬바이오)|금속(?!가공기계)|비금속|석화|페트로|카본(?!나노배터리)|합성|수지|LG화학|OCI|SK케미칼|한화솔루션(?!건설)|롯데케미칼|금호석화|효성화학|태광산업|KCC(?!건설)|고무|타이어소재|카본블랙|페놀|아크릴|스티렌|에틸렌|프로필렌|도금(?!기계)|강판|강재|봉강|선재|형강|후판|스테인리스|STS|내화재|정밀화학|정밀소재'),

# S13 기계·산업장비
('S13', r'기계(?!바이오|게임|금융)|산업장비|산업기계|산업용(?!가스에너지)|자동화(?!소프트웨어|SW|플랫폼)|로봇(?!소프트|AI플랫폼)|로보틱스|측정(?!IT|데이터)|제어(?!SW|IT|플랫폼)|전력기기|변압기|차단기|배전반|배전기|개폐기|발전기(?!재생에너지)|모터(?!IT)|인버터(?!에너지저장)|펌프(?!바이오|의료)|밸브(?!IT)|컨베이어|크레인|호이스트|공작기계|프레스|금형|사출(?!IT)|용접|절삭|연마|열교환기|냉동(?!식품)|에어컨(?!IT)|공조|냉각|에어컴프레서|압축기|진공장비|진공(?!의료바이오)|정공(?!무역)|정밀기계|정밀부품(?!반도체)|기공(?!IT)|산기(?!바이오)|설비(?!IT|소프트)|장치(?!의료|금융)|공업(?!화학|자동차|전자|조선|건설|식품|의약|광업)|산업(?!안전IT|건설|식품|광업|유통|금융|화학)|중공업(?!조선방산)|두산인프라코어|현대두산인프라코어|HD현대|두산밥캣|두산에너빌리티|LS일렉트릭|효성중공업|현대일렉트릭|에너지기계|보일러|터빈(?!배터리)|반도체장비(?!반도체자체)|디스플레이장비(?!디스플레이자체)|이차전지장비(?!전지자체)|배터리장비(?!배터리자체)|CNC|유압|공압|계측|계량|검사기(?!의료)|측정기|한빛레이저|가온전선|LS전선|대한전선|일진전기(?!S01)|고려전선|한솔기계'),

# S14 미디어·콘텐츠·광고
('S14', r'미디어|콘텐츠(?!관리SW)|방송(?!장비기계)|영화|드라마|예능|애니메이션|웹툰|웹소설|음악|음반|공연|뮤지컬|광고(?!기술IT)|출판|신문(?!IT)|잡지|만화|스튜디오|프로덕션|엔터테인먼트|SM엔터|YG엔터|JYP엔터|하이브|카카오엔터|CJ ENM|SBS|MBC|티빙|왓챠|멀티플렉스|CGV|메가박스|롯데시네마|디지틀조선|조선비즈|조선미디어|이데일리|머니투데이|OTT|스트리밍(?!게임)|K-POP|아이돌|연예|제이콘텐트리|덱스터|위지윅|NEW|바른손|영상(?!의료)|광고대행'),

# S15 식품·농업·수산
('S15', r'식품|음료(?!화학)|주류|맥주|소주|막걸리|와인|라면|과자|제과|제빵|베이커리|커피|농업|농산물|농약|비료(?!화학)|사료(?!기계)|축산|양계|양돈|낙농|수산|해산물|어업|수협|김치|장류|조미|양념|분유|유제품|아이스크림|빙과|CJ제일제당|오뚜기|농심|삼양식품|롯데칠성|하이트진로|오비맥주|매일유업|남양유업|빙그레|한국야쿠르트|풀무원|동원F&B|사조|대상(?!홀딩스건설)|샘표|청정원|hy(?!브리드)|종자|씨앗|묘목|원예|어분|어유|가공식품|즉석식품|냉동식품|간편식|HMR|사료작물|배합사료|이지바이오|인바이오|NH농우바이오|팜스코|해태제과|크라운제과|동서식품|동원산업|사조산업|KT&G(?!IT)|담배|홍삼|인삼|영농|농협'),

# S16 운송·물류·여행
('S16', r'항공(?!우주방산|우주산업|전자부품)|해운|선사|운송(?!장비기계)|물류|택배|포워딩|항만|터미널(?!IT)|여행|관광|호텔|리조트|숙박|크루즈|유람선|대한항공|아시아나|제주항공|진에어|에어부산|에어서울|HMM|팬오션|대한해운|흥아해운|CJ대한통운|한진(?!그룹전체)|롯데글로벌로지스|한국공항|인천공항|부산항만|철도|버스(?!IT)|로지스|운수(?!전자)|콜드체인|냉장물류|티웨이|에어프레미아'),

# S17 생활소비재·패션
('S17', r'화장품|코스메틱|뷰티(?!테크IT)|스킨케어|향수|헤어케어|의류|패션|의복|섬유(?!산업재화학)|니트|캐주얼|스포츠의류|아웃도어|가방|신발|잡화|가구|인테리어소품(?!건설)|생활용품|청소|세탁|위생(?!시설건설)|기저귀|생리대|면도|구강|레저용품|캠핑|골프(?!장비기계)|낚시|반려동물|펫(?!케어기술)|LG생활건강|아모레퍼시픽|한국콜마|코스맥스|에이블씨엔씨|토니모리|애경산업|F&F|한섬|LF|코오롱인더스트리(?!화학)|영원무역|에프앤에프|더네이쳐홀딩스|코스메|뷰(?!노의료)'),

# S18 환경·폐기물·재활용
('S18', r'환경(?!IT|솔루션SW|서비스IT)|폐기물|재활용|리사이클|수처리|정수|하수|폐수|대기오염|토양오염|환경설비|탄소중립|ESG환경|인선이엔티|코엔텍|와이엔텍|EMC홀딩스|에코비트|에코매니지먼트|에코(?!프로배터리|에너지전지)|케이에코|쓰레기|폐플라스틱|고형연료|집진|스크러버|수질검사|환경측정'),

# S19 교육·전문서비스
('S19', r'교육(?!기술IT)|학원|이러닝|에듀|유아교육|직업훈련|어학|학습|연구(?!소재화학|바이오화학)|컨설팅(?!IT)|경영컨설팅|시험(?!장비기계)|검사(?!장비기계)|인증(?!보안IT)|법무|회계|감사|세무|인사(?!IT)|HR(?!IT)|헤드헌팅|아웃소싱|메가스터디|이투스|대성학원|하늘교육|정상JLS|에듀윌|엔이능률|YBM|파고다|청담러닝|한국교육|교원그룹|대교|웅진씽크빅|능률교육|법률|특허|지식재산|변리'),

# S03 금융·보험·투자 (스팩 제외 일반 금융)
('S03', r'금융(?!기계)|은행|저축은행|신탁(?!부동산개발)|증권|자산운용|투자(?!부동산|IT)|캐피탈|카드(?!게임)|파이낸스|손해보험|생명보험|화재보험|재보험|공제|신한지주|KB금융|하나금융|우리금융|BNK|DGB|JB금융|한국금융|메리츠|DB금융|교보|흥국|현대해상|동부화재|삼성화재|DB손보|NH투자|미래에셋|삼성증권|키움증권|대신증권|신영증권|유안타|유진투자|LS증권|카카오뱅크|카카오페이|케이뱅크|인터넷뱅크|지주(?!산업|에너지|건설)|홀딩스(?!에너지|건설|화학)'),

]

MANUAL_FIXES = {
    '072870': 'S19',  # 메가스터디
    '215200': 'S19',  # 메가스터디교육
    '033130': 'S14',  # 디지틀조선 (미디어)
    '353810': 'S15',  # 이지바이오 (사료/농업)
    '352940': 'S15',  # 인바이오 (농업)
    '054050': 'S15',  # NH농우바이오 (종자/농업)
}

def classify(name: str, code: str):
    if code in MANUAL_FIXES:
        return MANUAL_FIXES[code], '수동 오류 보정'
    if re.search(r'스팩|SPAC|호스팩', name, re.I):
        return 'S03', '스팩(SPAC) 자동 분류'
    for sec, pat in PATTERNS:
        if re.search(pat, name):
            return sec, '자동 패턴 매칭'
    return 'S20', '미매칭 산업재·기타제조'

# ────────────────────────────────────────────
# 데이터 로드
# ────────────────────────────────────────────
with open(r'D:\logofchoices\data\csv\all_stocks.json', encoding='utf-8') as f:
    raw = json.load(f)

df = pd.DataFrame(raw, columns=['n','c','m','s'])
df.columns = ['종목명','종목코드','시장','기존섹터코드']
df['종목코드'] = df['종목코드'].astype(str).str.zfill(6)
df['기존섹터명'] = df['기존섹터코드'].map(lambda c: SECTOR_META.get(c, ('기타','기타'))[0])

# 재분류
results = df.apply(lambda r: classify(r['종목명'], r['종목코드']), axis=1)
df['섹터코드']  = results.map(lambda x: x[0])
df['섹터명']    = df['섹터코드'].map(lambda c: SECTOR_META[c][0])
df['분류근거']  = results.map(lambda x: x[1])
df['매칭방식']  = df['분류근거'].map(
    lambda x: '수동보정' if '수동' in x else ('스팩' if '스팩' in x else '키워드'))
df['검토신뢰도'] = df.apply(
    lambda r: '높음' if r['매칭방식'] in ('수동보정','스팩')
              else ('높음' if r['섹터코드'] != 'S20' else '중간'), axis=1)
df['추가검토필요'] = df['검토신뢰도'].map(lambda x: True if x == '중간' else False)

df = df.sort_values(['섹터코드','시장','종목명']).reset_index(drop=True)
FINAL_COLS = ['종목명','종목코드','시장','섹터코드','섹터명',
              '기존섹터코드','기존섹터명','분류근거','매칭방식','검토신뢰도','추가검토필요']
df_out = df[FINAL_COLS]

# ────────────────────────────────────────────
# CSV 저장
# ────────────────────────────────────────────
csv_path = r'D:\logofchoices\data\csv\stocks_sector_final.csv'
df_out.to_csv(csv_path, index=False, encoding='utf-8-sig')
print(f'CSV: {csv_path}')

sec_cnt = Counter(df_out['섹터코드'])
tot = len(df_out)
print(f'\n[Total] {tot:,}  [S99] {sec_cnt.get("S99",0)}')
print(f'{"Code":<5} {"Name":<26} {"Count":>6} {"Pct":>6}')
print('-' * 48)
for code in [f'S{i:02d}' for i in range(1,21)]:
    n = sec_cnt.get(code, 0)
    name = SECTOR_META[code][0]
    print(f'{code:<5} {name:<26} {n:>6,} {n/tot*100:>5.1f}%')
print(f'{"Total":<32} {tot:>6,} 100.0%')

review_cnt  = df_out['추가검토필요'].sum()
changed_cnt = (df_out['섹터코드'] != df_out['기존섹터코드']).sum()
manual_cnt  = (df_out['분류근거'] == '수동 오류 보정').sum()
print(f'\n[Changed] {changed_cnt:,}  [Manual] {manual_cnt}  [ReviewNeeded] {review_cnt:,}')

# ────────────────────────────────────────────
# 엑셀 생성
# ────────────────────────────────────────────
HEADER_FONT  = Font(name='맑은 고딕', bold=True, color='FFFFFF', size=10)
HEADER_FILL  = PatternFill(fill_type='solid', start_color='1F4E79')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
BODY_ALIGN   = Alignment(vertical='center')
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
thin = Side(style='thin', color='CCCCCC')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def set_header(ws, cols, row=1):
    for c, title in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=title)
        cell.font, cell.fill, cell.alignment, cell.border = HEADER_FONT, HEADER_FILL, HEADER_ALIGN, BORDER

def col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

wb = Workbook()
wb.remove(wb.active)

# 시트1: 섹터정의
ws1 = wb.create_sheet('섹터정의')
ws1.row_dimensions[1].height = 28
set_header(ws1, ['섹터코드','섹터명','포함범위','종목수'])
col_widths(ws1, [10, 24, 50, 10])
for i, (code, (name, scope)) in enumerate(SECTOR_META.items(), 2):
    fill = PatternFill(fill_type='solid', start_color=SECTOR_COLORS[code])
    for c, val in enumerate([code, name, scope, sec_cnt.get(code, 0)], 1):
        cell = ws1.cell(row=i, column=c, value=val)
        cell.fill, cell.border = fill, BORDER
        cell.font = Font(name='맑은 고딕', size=9)
        cell.alignment = CENTER_ALIGN if c in (1, 4) else BODY_ALIGN
ws1.freeze_panes = 'A2'

# 시트2: 전체분류
ws2 = wb.create_sheet('전체분류')
ws2.row_dimensions[1].height = 28
cols2 = ['종목명','종목코드','시장','섹터코드','섹터명','기존섹터코드','기존섹터명','분류근거','매칭방식','검토신뢰도','추가검토필요']
set_header(ws2, cols2)
col_widths(ws2, [22, 12, 10, 10, 22, 10, 22, 20, 12, 12, 14])
REVIEW_FILL = PatternFill(fill_type='solid', start_color='FFEB9C')
for i, row in enumerate(df_out.itertuples(index=False), 2):
    sec = row.섹터코드
    base_fill = PatternFill(fill_type='solid', start_color=SECTOR_COLORS.get(sec,'F5F5F5'))
    vals = [row.종목명, row.종목코드, row.시장, row.섹터코드, row.섹터명,
            row.기존섹터코드, row.기존섹터명, row.분류근거, row.매칭방식,
            row.검토신뢰도, str(row.추가검토필요)]
    for c, val in enumerate(vals, 1):
        cell = ws2.cell(row=i, column=c, value=val)
        cell.fill = REVIEW_FILL if row.추가검토필요 else base_fill
        cell.border, cell.font = BORDER, Font(name='맑은 고딕', size=9)
        cell.alignment = CENTER_ALIGN if c in (2,3,4,6,9,10,11) else BODY_ALIGN
ws2.freeze_panes = 'A2'
ws2.auto_filter.ref = f'A1:{get_column_letter(len(cols2))}1'

# 시트3: 변경내역
ws3 = wb.create_sheet('변경내역')
ws3.row_dimensions[1].height = 28
cols3 = ['종목명','종목코드','시장','기존섹터코드','기존섹터명','섹터코드','섹터명','분류근거','검토신뢰도']
set_header(ws3, cols3)
col_widths(ws3, [22, 12, 10, 10, 22, 10, 22, 20, 12])
changed = df_out[df_out['섹터코드'] != df_out['기존섹터코드']].copy()
MANUAL_FILL = PatternFill(fill_type='solid', start_color='FFC7CE')
for i, row in enumerate(changed.itertuples(index=False), 2):
    is_manual = row.분류근거 == '수동 오류 보정'
    sec = row.섹터코드
    base_fill = MANUAL_FILL if is_manual else PatternFill(fill_type='solid', start_color=SECTOR_COLORS.get(sec,'F5F5F5'))
    vals = [row.종목명, row.종목코드, row.시장, row.기존섹터코드, row.기존섹터명,
            row.섹터코드, row.섹터명, row.분류근거, row.검토신뢰도]
    for c, val in enumerate(vals, 1):
        cell = ws3.cell(row=i, column=c, value=val)
        cell.fill, cell.border, cell.font = base_fill, BORDER, Font(name='맑은 고딕', size=9)
        cell.alignment = CENTER_ALIGN if c in (2,3,4,6,9) else BODY_ALIGN
ws3.freeze_panes = 'A2'
ws3.auto_filter.ref = f'A1:{get_column_letter(len(cols3))}1'

# 시트4: 추가검토권장
ws4 = wb.create_sheet('추가검토권장')
ws4.row_dimensions[1].height = 28
cols4 = ['종목명','종목코드','시장','섹터코드','섹터명','기존섹터코드','분류근거','매칭방식','검토신뢰도']
set_header(ws4, cols4)
col_widths(ws4, [22, 12, 10, 10, 22, 10, 20, 12, 12])
review_df = df_out[df_out['추가검토필요'] == True].copy()
for i, row in enumerate(review_df.itertuples(index=False), 2):
    sec = row.섹터코드
    fill = PatternFill(fill_type='solid', start_color=SECTOR_COLORS.get(sec,'F5F5F5'))
    vals = [row.종목명, row.종목코드, row.시장, row.섹터코드, row.섹터명,
            row.기존섹터코드, row.분류근거, row.매칭방식, row.검토신뢰도]
    for c, val in enumerate(vals, 1):
        cell = ws4.cell(row=i, column=c, value=val)
        cell.fill, cell.border, cell.font = fill, BORDER, Font(name='맑은 고딕', size=9)
        cell.alignment = CENTER_ALIGN if c in (2,3,4,6,9) else BODY_ALIGN
ws4.freeze_panes = 'A2'
ws4.auto_filter.ref = f'A1:{get_column_letter(len(cols4))}1'

# 시트5: 섹터별요약
ws5 = wb.create_sheet('섹터별요약')
ws5.row_dimensions[1].height = 28
cols5 = ['섹터코드','섹터명','KOSPI','KOSDAQ','합계','비율(%)']
set_header(ws5, cols5)
col_widths(ws5, [10, 26, 12, 12, 12, 12])
kospi_cnt  = df_out[df_out['시장']=='KOSPI'].groupby('섹터코드')['종목코드'].count()
kosdaq_cnt = df_out[df_out['시장']=='KOSDAQ'].groupby('섹터코드')['종목코드'].count()
for i, code in enumerate([f'S{j:02d}' for j in range(1,21)], 2):
    name  = SECTOR_META[code][0]
    kp    = int(kospi_cnt.get(code, 0))
    kq    = int(kosdaq_cnt.get(code, 0))
    total = kp + kq
    pct   = round(total / tot * 100, 1)
    fill  = PatternFill(fill_type='solid', start_color=SECTOR_COLORS[code])
    for c, val in enumerate([code, name, kp, kq, total, pct], 1):
        cell = ws5.cell(row=i, column=c, value=val)
        cell.fill, cell.border, cell.font = fill, BORDER, Font(name='맑은 고딕', size=9)
        cell.alignment = CENTER_ALIGN if c != 2 else BODY_ALIGN
bold_fill = PatternFill(fill_type='solid', start_color='1F4E79')
tp = int(kospi_cnt.sum()); tq = int(kosdaq_cnt.sum())
for c, val in enumerate(['합계', '합계', tp, tq, tot, 100.0], 1):
    cell = ws5.cell(row=22, column=c, value=val)
    cell.fill, cell.border = bold_fill, BORDER
    cell.font = Font(name='맑은 고딕', bold=True, color='FFFFFF', size=10)
    cell.alignment = CENTER_ALIGN
ws5.freeze_panes = 'A2'

xlsx_path = r'D:\logofchoices\data\csv\stocks_sector_final.xlsx'
wb.save(xlsx_path)
print(f'\nExcel: {xlsx_path}')

# ────────────────────────────────────────────
# 검증
# ────────────────────────────────────────────
print('\n---- Verification ----')
checks = {
    f'Total stocks = {tot}': tot == 2766,
    'S99 remaining = 0': sec_cnt.get('S99', 0) == 0,
    'Only S01~S20': all(c in SECTOR_META for c in df_out['섹터코드'].unique()),
}
for msg, (code, exp) in {
    'Megastudy(072870)->S19': ('072870', 'S19'),
    'MegastudyEdu(215200)->S19': ('215200', 'S19'),
    'DigitalJosun(033130)->S14': ('033130', 'S14'),
    'EasybioFeed(353810)->S15': ('353810', 'S15'),
    'NHCropSci(054050)->S15': ('054050', 'S15'),
}.items():
    mask = df_out['종목코드'] == code
    checks[msg] = (df_out.loc[mask, '섹터코드'].iloc[0] == exp) if mask.any() else False

all_ok = True
for msg, ok in checks.items():
    if not ok: all_ok = False
    print(f'  [{"OK" if ok else "NG"}] {msg}')
print(f'\n{">>> All passed!" if all_ok else ">>> Some failed!"}')
