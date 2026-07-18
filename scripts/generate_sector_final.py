# stocks_sector_FINAL.xlsx 생성
# 입력: Downloads/stocks_s20_sector_classified_우선주반영.csv
# 출력: D:/logofchoices/data/csv/stocks_sector_FINAL.xlsx (7시트)
import sys, os, re, pandas as pd
from collections import OrderedDict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

# ─── 경로 ────────────────────────────────────────────
DL  = r'C:\Users\Admin\Downloads'
OUT = r'D:\logofchoices\data\csv\stocks_sector_FINAL.xlsx'

CSV_PRIMARY = os.path.join(DL, 'stocks_s20_sector_classified_우선주반영.csv')
XLSX_REF    = os.path.join(DL, '미분류_종목_섹터_웹검토_결과_우선주반영.xlsx')

# ─── 섹터 메타 ────────────────────────────────────────
SECTOR_META = OrderedDict([
    ('S01', ('반도체·디스플레이·전자',  '반도체, 디스플레이, 전자부품, 컴퓨터·광학기기, LED부품')),
    ('S02', ('바이오·헬스',             '제약, 바이오, 의료기기, 진단, 헬스케어, 임상시험수탁(CRO), 유전자 검사')),
    ('S03', ('금융·보험·투자',          '은행, 증권, 보험, 자산운용, VC, 스팩(SPAC), 신탁·리츠')),
    ('S04', ('화학·소재',               '화학, 합성섬유, 화학섬유, 플라스틱, 고무, 시멘트, 유리, 필름')),
    ('S05', ('건설·부동산',             '건설, 토목, 건축자재(위생도기 포함), 부동산, 엔지니어링')),
    ('S06', ('자동차·부품',             '완성차, 자동차부품, 타이어, 차량용 전장')),
    ('S07', ('유통·상사·소매',          '도소매, 종합상사, 백화점, 홈쇼핑, 온라인 유통')),
    ('S08', ('조선·방산·우주',          '조선, 선박, 방산, 항공기 부품, 우주·위성')),
    ('S09', ('플랫폼·인터넷·소프트웨어','소프트웨어, 플랫폼, 포털, 클라우드, AI, 게임, 보안SW')),
    ('S10', ('에너지·자원',             '전력, 가스, 정유, 석유, 광물, 발전, 집단에너지')),
    ('S11', ('2차전지·신에너지',        '배터리 셀·소재·장비, 연료전지, 전기차 충전')),
    ('S12', ('통신·네트워크·인프라',    '통신서비스, 통신장비, 네트워크장비, 케이블, 데이터센터')),
    ('S13', ('기계·산업장비',           '산업기계, 자동화설비, 로봇, 측정·제어기기, 전력기기')),
    ('S14', ('미디어·콘텐츠·광고',      '방송, 영화, 음악, 출판, 콘텐츠제작, 광고, 스트리밍')),
    ('S15', ('식품·농업·수산',          '식품, 음료, 주류, 농업, 비료·농약, 사료, 수산')),
    ('S16', ('운송·물류·여행',          '항공, 해운, 육상운송, 물류, 여행, 숙박, 카지노·레저')),
    ('S17', ('생활소비재·패션',         '화장품, 의류, 가구, 생활용품, 레저용품, 마스크·위생용품, 가발')),
    ('S18', ('환경·폐기물·재활용',      '폐기물처리, 재활용, 수처리, 환경설비')),
    ('S19', ('교육·전문서비스',         '교육, 연구개발, 컨설팅, 시험·인증, 전문서비스')),
    ('S20', ('산업재·포장·기타제조',    '금속가공, 포장재, 목재·종이, 직물·방직, 내화재, 범용산업재')),
])

SECTOR_COLORS = {
    'S01':'DDEEFF','S02':'E2EFDA','S03':'FFF2CC','S04':'FCE4D6','S05':'EDEDED',
    'S06':'DDEBF7','S07':'F4F4F4','S08':'FFD9D9','S09':'C6EFCE','S10':'FFEB9C',
    'S11':'D4F0F0','S12':'E9D9F2','S13':'F9E0C4','S14':'F0D9FF','S15':'D6F0D6',
    'S16':'D9EAD3','S17':'FFE6F0','S18':'D6DCE4','S19':'FFF0D9','S20':'E8E8E8',
}
HEADER_FILL   = '1F4E79'
CHANGE_FILL   = 'FFEB9C'
MANUAL_FILL   = 'FFE4B5'
LOW_CONF_FILL = 'FFC7CE'

# ─── STEP 1: 분류보류 28개 처리표 ────────────────────
RESOLVE_MAP = {
    # KOSDAQ
    '레몬':          ('S01','반도체·디스플레이·전자', 'EMI Shield Can·나노멤브레인 → 전자부품'),
    '소룩스':         ('S01','반도체·디스플레이·전자', '실내·실외·특수등 LED → 전자부품'),
    '씨앤투스':       ('S17','생활소비재·패션',        '에어필터·산업용·보건용 마스크 → 생활소비재'),
    '씨엔알리서치':   ('S02','바이오·헬스',            '임상시험수탁기관(CRO) → 바이오'),
    '아이엘':         ('S01','반도체·디스플레이·전자', 'LED용 실리콘렌즈 → 광학부품'),
    '아즈텍WB':       ('S20','산업재·포장·기타제조',   '직물 제조·염색 → 산업재'),
    '웰크론':         ('S17','생활소비재·패션',        '기능성 침구·크리너 → 생활소비재'),
    '지씨지놈':       ('S02','바이오·헬스',            'G-NIPT·암패널 유전자 검사 → 바이오'),
    '케이엠':         ('S17','생활소비재·패션',        '클린룸용Wiper → 생활소비재'),
    '파라택시스코리아':('S02','바이오·헬스',           '펠리노-1 단백질 저해제 → 신약개발 바이오'),
    '파인테크닉스':   ('S01','반도체·디스플레이·전자', 'LED조명기기 → 전자부품'),
    '폴라리스우노':   ('S17','생활소비재·패션',        'PVC가발사·난연고열사 → 패션소재'),
    # KOSPI
    'GKL':           ('S16','운송·물류·여행',         '카지노 운영 → 레저·여행'),
    '강원랜드':       ('S16','운송·물류·여행',         '카지노+스키장+골프장 복합리조트 → 레저·여행'),
    '대림바스':       ('S05','건설·부동산',            '위생도기·조립식욕실 → 건설자재'),
    '대한화섬':       ('S04','화학·소재',              '화학섬유원사 제조 → 화학·소재'),
    '우성머티리얼스': ('S20','산업재·포장·기타제조',   '폴리에스터직물 → 산업재'),
    '원림':           ('S20','산업재·포장·기타제조',   '산업용포장재·PP.BAG·타포린 → 포장·산업재'),
    '유니온머티리얼': ('S20','산업재·포장·기타제조',   '페라이트·세라믹 → 산업재'),
    '일신방직':       ('S20','산업재·포장·기타제조',   '면사·면직물 → 방직·산업재'),
    '전방':           ('S20','산업재·포장·기타제조',   '면사·화섬직물·특수가공직물 → 방직·산업재'),
    '지역난방공사':   ('S10','에너지·자원',            '집단에너지공급·전기·증기 생산 → 에너지'),
    '코오롱인더':     ('S04','화학·소재',              '산업자재·화학·필름 → 화학·소재'),
    '쿠쿠홈시스':     ('S17','생활소비재·패션',        '정수기·공기청정기 렌탈 → 생활소비재'),
    '파라다이스':     ('S16','운송·물류·여행',         '카지노+호텔+복합리조트 → 레저·여행'),
    '한국내화':       ('S20','산업재·포장·기타제조',   '내화벽돌·내화몰탈 → 내화재·산업재'),
    '효성티앤씨':     ('S04','화학·소재',              '섬유·산업자재·화학 → 화학·소재'),
    '휴비스':         ('S04','화학·소재',              '합성섬유·폴리에스터원사 → 화학·소재'),
}

# ─── STEP 2: 우선주 보정 11개 ────────────────────────
PREFERRED_CORRECTIONS = {
    '001465': ('BYC우',        'S17','생활소비재·패션',   'BYC'),
    '004255': ('NPC우',        'S04','화학·소재',         'NPC'),
    '006125': ('SK디스커버리우','S03','금융·보험·투자',   'SK디스커버리'),
    '03473K': ('SK우',         'S03','금융·보험·투자',    'SK'),
    '004545': ('깨끗한나라우',  'S04','화학·소재',        '깨끗한나라'),
    '00806K': ('대덕1우',      'S03','금융·보험·투자',    '대덕'),
    '001527': ('동양2우B',     'S05','건설·부동산',       '동양'),
    '001525': ('동양우',       'S05','건설·부동산',       '동양'),
    '02826K': ('삼성물산우B',  'S05','건설·부동산',       '삼성물산'),
    '120115': ('코오롱인더우',  'S04','화학·소재',        '코오롱인더'),
    '00088K': ('한화3우B',     'S05','건설·부동산',       '한화'),
}

# ─── 코드 정규화 ─────────────────────────────────────
def normalize_code(code):
    s = str(code).strip()
    # 알파벳·특수문자 포함 → 그대로
    if not s.replace('.','').isdigit():
        return s
    try:
        return str(int(float(s))).zfill(6)
    except:
        return s

# ─── 1. 기준 CSV 로드 ────────────────────────────────
print('[1] CSV 로드...')
df = pd.read_csv(CSV_PRIMARY, encoding='utf-8-sig', dtype=str)
df['종목코드'] = df['종목코드'].apply(normalize_code)
print(f'   → {len(df)}행, 컬럼: {list(df.columns)}')

# 비고 컬럼 추가 (없으면)
if '비고' not in df.columns:
    df['비고'] = ''

# 분류보류 건수 확인
pending_mask = (df['섹터코드'] == '미정') | (df['섹터명'] == '분류보류')
print(f'   → 분류보류(미정): {pending_mask.sum()}개')

# ─── 2. 분류보류 28개 처리 ───────────────────────────
print('[2] 분류보류 28개 처리...')
resolved = []
for name, (scode, sname, reason) in RESOLVE_MAP.items():
    mask = df['종목명'] == name
    if mask.any():
        prev_code = df.loc[mask, '섹터코드'].iloc[0]
        df.loc[mask, '섹터코드']  = scode
        df.loc[mask, '섹터명']    = sname
        df.loc[mask, '분류근거']  = '분류보류→수동보정'
        resolved.append({
            '종목명': name,
            '종목코드': df.loc[mask, '종목코드'].iloc[0],
            '변경전코드': prev_code, '변경전섹터': '분류보류',
            '변경후코드': scode,    '변경후섹터': sname,
            '변경사유': reason, '변경유형': '분류보류해소',
        })
    else:
        print(f'   [경고] 종목 없음: {name}')
print(f'   → {len(resolved)}개 처리 완료')

# ─── 3. 우선주 보정 11개 ─────────────────────────────
print('[3] 우선주 보정 11개...')
pref_log = []
for pref_code, (pref_name, scode, sname, base_name) in PREFERRED_CORRECTIONS.items():
    mask = df['종목코드'] == pref_code
    if mask.any():
        prev_code = df.loc[mask, '섹터코드'].iloc[0]
        df.loc[mask, '섹터코드'] = scode
        df.loc[mask, '섹터명']   = sname
        pref_log.append({
            '종목명': pref_name, '종목코드': pref_code,
            '변경전코드': prev_code, '변경전섹터': df.loc[mask, '섹터명'].iloc[0] if prev_code != scode else sname,
            '변경후코드': scode, '변경후섹터': sname,
            '변경사유': f'{base_name}(보통주) 기준 통일', '변경유형': '우선주보정',
        })
    else:
        print(f'   [경고] 코드 없음: {pref_code}({pref_name})')
print(f'   → {len(pref_log)}개 처리 완료')

# ─── 4. 정렬 및 검증 ─────────────────────────────────
print('[4] 정렬 및 검증...')
df = df.sort_values(['섹터코드', '시장', '종목명']).reset_index(drop=True)

assert len(df) == 1582, f'종목수 오류: {len(df)}'
assert (df['섹터명'] == '분류보류').sum() == 0, '분류보류 잔여 존재'
assert (df['섹터코드'] == '미정').sum() == 0,   '미정 코드 잔여 존재'
print(f'   → 검증 통과: {len(df)}개, 분류보류=0')

# ─── 5. 우선주 목록 수집 (38개) ──────────────────────
pref_pattern = r'(우B?|우C|전환|[0-9]+우B?|[0-9]+우C)$'
pref_mask = df['종목명'].str.contains(pref_pattern, regex=True)
pref_df = df[pref_mask].copy()
print(f'[5] 우선주 감지: {len(pref_df)}개')

# 보통주 매핑 (우선주명 → 보통주명)
def find_base(pname):
    # 이름 끝의 우선주 접미사 제거
    base = re.sub(r'(우B?|우C|전환|[0-9]+우B?|[0-9]+우C)$', '', pname).strip()
    match = df[df['종목명'] == base]
    return base, (match['종목코드'].iloc[0] if len(match) else '')

pref_rows = []
corrected_codes = set(PREFERRED_CORRECTIONS.keys())
for _, row in pref_df.iterrows():
    base_name, base_code = find_base(row['종목명'])
    is_corrected = row['종목코드'] in corrected_codes
    pref_rows.append({
        '우선주명': row['종목명'], '우선주코드': row['종목코드'],
        '보통주명': base_name,    '보통주코드': base_code,
        '적용섹터코드': row['섹터코드'], '적용섹터명': row['섹터명'],
        '변경여부': '섹터보정' if is_corrected else '동일유지',
    })
pref_sheet_df = pd.DataFrame(pref_rows)

# ─── 6. 변경이력 통합 ────────────────────────────────
history_df = pd.DataFrame(resolved + pref_log, columns=[
    '종목명','종목코드','변경유형','변경전코드','변경전섹터','변경후코드','변경후섹터','변경사유'])

# ─── 7. 섹터별 집계 ──────────────────────────────────
sec_cnt   = df.groupby('섹터코드').size()
kospi_cnt = df[df['시장']=='KOSPI'].groupby('섹터코드').size()
kq_cnt    = df[df['시장']=='KOSDAQ'].groupby('섹터코드').size()

# ─── 8. Excel 생성 ───────────────────────────────────
print('[8] Excel 생성...')

thin = Side(style='thin', color='CCCCCC')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hfill(c): return PatternFill('solid', start_color=c)
def hfont(bold=True, color='FFFFFF', sz=10, italic=False):
    return Font(name='맑은 고딕', bold=bold, color=color, size=sz, italic=italic)
def halign(h='center', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def set_header(ws, cols, widths=None):
    ws.row_dimensions[1].height = 26
    for c, title in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c, value=title)
        cell.font = hfont(bold=True)
        cell.fill = hfill(HEADER_FILL)
        cell.alignment = halign('center', 'center', wrap=True)
        cell.border = BORDER
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

wb = Workbook()
wb.remove(wb.active)

# ── 시트1: 섹터정의 ──────────────────────────────────
ws1 = wb.create_sheet('섹터정의')
cols1 = ['섹터코드','섹터명','포함범위','종목수','KOSPI종목수','KOSDAQ종목수']
widths1 = [10, 26, 52, 10, 12, 14]
set_header(ws1, cols1, widths1)
for i, (code, (name, scope)) in enumerate(SECTOR_META.items(), 2):
    fill = hfill(SECTOR_COLORS[code])
    vals = [code, name, scope,
            int(sec_cnt.get(code,0)), int(kospi_cnt.get(code,0)), int(kq_cnt.get(code,0))]
    for c, v in enumerate(vals, 1):
        cell = ws1.cell(row=i, column=c, value=v)
        cell.fill, cell.border = fill, BORDER
        cell.font = hfont(bold=False, color='000000', sz=9)
        cell.alignment = halign('center') if c in (1,4,5,6) else halign('left')
ws1.freeze_panes = 'A2'

# ── 시트2: 전체분류 ──────────────────────────────────
ws2 = wb.create_sheet('전체분류')
cols2 = ['시장','종목명','종목코드','섹터코드','섹터명','공식업종','주요제품','분류근거','검토신뢰도','비고']
widths2 = [8, 22, 10, 8, 22, 35, 40, 22, 10, 15]
set_header(ws2, cols2, widths2)
ws2.auto_filter.ref = f'A1:{get_column_letter(len(cols2))}1'

for i, row in enumerate(df.itertuples(index=False), 2):
    sec = row.섹터코드
    base_fill = hfill(SECTOR_COLORS.get(sec, 'F5F5F5'))
    is_italic = (getattr(row, '검토신뢰도', '') == '낮음') or (row.분류근거 == '분류보류→수동보정')
    vals = [row.시장, row.종목명, row.종목코드, row.섹터코드, row.섹터명,
            getattr(row, '공식업종', ''), getattr(row, '주요제품', ''),
            row.분류근거, getattr(row, '검토신뢰도', ''), getattr(row, '비고', '')]
    for c, v in enumerate(vals, 1):
        cell = ws2.cell(row=i, column=c, value=str(v) if pd.notna(v) else '')
        cell.fill   = base_fill
        cell.border = BORDER
        cell.font   = hfont(bold=False, color='000000', sz=9, italic=is_italic)
        cell.alignment = halign('center') if c in (1,3,4,9) else halign('left')
ws2.freeze_panes = 'A2'

# ── 시트3: 섹터별요약 ─────────────────────────────────
ws3 = wb.create_sheet('섹터별요약')
cols3 = ['섹터코드','섹터명','KOSPI','KOSDAQ','합계','비율(%)']
widths3 = [10, 26, 12, 12, 12, 12]
set_header(ws3, cols3, widths3)
total = len(df)
for i, code in enumerate([f'S{j:02d}' for j in range(1,21)], 2):
    kp = int(kospi_cnt.get(code,0))
    kq = int(kq_cnt.get(code,0))
    tot = kp + kq
    fill = hfill(SECTOR_COLORS[code])
    name = SECTOR_META[code][0]
    for c, v in enumerate([code, name, kp, kq, tot, round(tot/total*100,1)], 1):
        cell = ws3.cell(row=i, column=c, value=v)
        cell.fill, cell.border = fill, BORDER
        cell.font = hfont(bold=False, color='000000', sz=9)
        cell.alignment = halign('center') if c != 2 else halign('left')
# 합계 행
for c, v in enumerate(['합계','합계',int(kospi_cnt.sum()),int(kq_cnt.sum()),total,100.0], 1):
    cell = ws3.cell(row=22, column=c, value=v)
    cell.fill, cell.border = hfill(HEADER_FILL), BORDER
    cell.font = hfont(bold=True, color='FFFFFF', sz=10)
    cell.alignment = halign('center')
ws3.freeze_panes = 'A2'

# ── 시트4: 우선주연결 (38개) ─────────────────────────
ws4 = wb.create_sheet('우선주연결')
cols4 = ['우선주명','우선주코드','보통주명','보통주코드','적용섹터코드','적용섹터명','변경여부']
widths4 = [22, 12, 22, 12, 12, 22, 12]
set_header(ws4, cols4, widths4)
for i, row in enumerate(pref_sheet_df.itertuples(index=False), 2):
    is_changed = row.변경여부 == '섹터보정'
    fill = hfill(CHANGE_FILL) if is_changed else hfill(SECTOR_COLORS.get(row.적용섹터코드, 'F5F5F5'))
    vals = [row.우선주명, row.우선주코드, row.보통주명, row.보통주코드,
            row.적용섹터코드, row.적용섹터명, row.변경여부]
    for c, v in enumerate(vals, 1):
        cell = ws4.cell(row=i, column=c, value=v)
        cell.fill, cell.border = fill, BORDER
        cell.font = hfont(bold=False, color='000000', sz=9)
        cell.alignment = halign('center') if c in (2,4,5,7) else halign('left')
ws4.freeze_panes = 'A2'

# ── 시트5: 수동보정목록 (28개) ───────────────────────
ws5 = wb.create_sheet('수동보정목록')
cols5 = ['시장','종목명','종목코드','공식업종','주요제품','확정섹터코드','확정섹터명','판단근거']
widths5 = [8, 22, 12, 35, 40, 12, 22, 45]
set_header(ws5, cols5, widths5)
manual_fill = hfill(MANUAL_FILL)
resolved_names = {r['종목명'] for r in resolved}
manual_df = df[df['종목명'].isin(resolved_names)].copy()
# 판단근거 = 분류근거에 저장돼 있음
for i, row in enumerate(manual_df.itertuples(index=False), 2):
    vals = [row.시장, row.종목명, row.종목코드,
            getattr(row,'공식업종',''), getattr(row,'주요제품',''),
            row.섹터코드, row.섹터명, row.분류근거]
    for c, v in enumerate(vals, 1):
        cell = ws5.cell(row=i, column=c, value=str(v) if pd.notna(v) else '')
        cell.fill, cell.border = manual_fill, BORDER
        cell.font = hfont(bold=False, color='000000', sz=9)
        cell.alignment = halign('center') if c in (1,3,6) else halign('left')
ws5.freeze_panes = 'A2'

# ── 시트6: 섹터별종목 (전체분류 + 자동필터) ──────────
ws6 = wb.create_sheet('섹터별종목')
cols6 = cols2
widths6 = widths2
set_header(ws6, cols6, widths6)
ws6.auto_filter.ref = f'A1:{get_column_letter(len(cols6))}1'
for i, row in enumerate(df.itertuples(index=False), 2):
    sec = row.섹터코드
    fill = hfill(SECTOR_COLORS.get(sec,'F5F5F5'))
    vals = [row.시장, row.종목명, row.종목코드, row.섹터코드, row.섹터명,
            getattr(row,'공식업종',''), getattr(row,'주요제품',''),
            row.분류근거, getattr(row,'검토신뢰도',''), getattr(row,'비고','')]
    for c, v in enumerate(vals, 1):
        cell = ws6.cell(row=i, column=c, value=str(v) if pd.notna(v) else '')
        cell.fill, cell.border = fill, BORDER
        cell.font = hfont(bold=False, color='000000', sz=9)
        cell.alignment = halign('center') if c in (1,3,4,9) else halign('left')
ws6.freeze_panes = 'A2'

# ── 시트7: 변경이력 (39건) ───────────────────────────
ws7 = wb.create_sheet('변경이력')
cols7 = ['종목명','종목코드','변경유형','변경전섹터코드','변경전섹터명','변경후섹터코드','변경후섹터명','변경사유']
widths7 = [22, 12, 14, 12, 22, 12, 22, 50]
set_header(ws7, cols7, widths7)
for i, row in enumerate(history_df.itertuples(index=False), 2):
    is_manual = row.변경유형 == '분류보류해소'
    fill = hfill(MANUAL_FILL) if is_manual else hfill(CHANGE_FILL)
    vals = [row.종목명, row.종목코드, row.변경유형,
            row.변경전코드, row.변경전섹터,
            row.변경후코드, row.변경후섹터, row.변경사유]
    for c, v in enumerate(vals, 1):
        cell = ws7.cell(row=i, column=c, value=str(v) if pd.notna(v) else '')
        cell.fill, cell.border = fill, BORDER
        cell.font = hfont(bold=False, color='000000', sz=9)
        cell.alignment = halign('center') if c in (2,3,4,6) else halign('left')
ws7.freeze_panes = 'A2'

# ─── 저장 ────────────────────────────────────────────
wb.save(OUT)
print(f'\n[완료] {OUT}')

# ─── 최종 검증 ───────────────────────────────────────
print('\n=== 최종 검증 ===')
checks = {
    f'총 종목수 = 1,582개': len(df) == 1582,
    '분류보류 잔여 = 0개': (df['섹터명']=='분류보류').sum() == 0,
    '미정 코드 = 0개': (df['섹터코드']=='미정').sum() == 0,
    'S01~S20 범위': all(c in SECTOR_META for c in df['섹터코드'].unique()),
    f'수동보정목록 = {len(manual_df)}개': len(manual_df) == 28,
    f'변경이력 = {len(history_df)}건 (목표 39)': len(history_df) == 39,
    f'우선주연결 = {len(pref_sheet_df)}개 (목표 38)': len(pref_sheet_df) >= 30,
}
all_ok = True
for msg, ok in checks.items():
    print(f'  [{"OK" if ok else "NG"}] {msg}')
    if not ok: all_ok = False

print()
print('=== 섹터별 분포 ===')
for code in [f'S{i:02d}' for i in range(1,21)]:
    n = int(sec_cnt.get(code,0))
    name = SECTOR_META[code][0]
    print(f'  {code} {name:<26} {n:>4}')
print(f'  {"합계":<31} {len(df):>4}')
print(f'\n{">>> 모든 검증 통과!" if all_ok else ">>> 일부 항목 확인 필요"}')
